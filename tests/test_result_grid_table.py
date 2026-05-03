from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from scripts.result_grid_table import (
    build_result_grid_score_table,
    export_result_grid_score_excel,
    list_result_grid_csvs,
)


def test_build_result_grid_score_table_excludes_all_csv(tmp_path):
    pd.DataFrame(
        [
            {"dataset": "mnist", "F_type": "umap", "G_type": "linear", "score_mean": 0.1},
            {"dataset": "mnist", "F_type": "krr", "G_type": "linear", "score_mean": 0.2},
        ]
    ).to_csv(tmp_path / "result_grid_mnist.csv", index=False)
    pd.DataFrame(
        [{"dataset": "all", "F_type": "ignored", "G_type": "ignored", "score_mean": 9.9}]
    ).to_csv(tmp_path / "result_grid_all.csv", index=False)

    assert [path.name for path in list_result_grid_csvs(tmp_path)] == ["result_grid_mnist.csv"]

    table = build_result_grid_score_table(tmp_path)

    assert table.loc["linear", ("mnist", "umap")] == 0.1
    assert table.loc["linear", ("mnist", "krr")] == 0.2
    assert ("all", "ignored") not in table.columns


def test_export_result_grid_score_excel_layout(tmp_path):
    pd.DataFrame(
        [
            {"dataset": "mnist", "F_type": "umap", "G_type": "linear", "score_mean": 0.1},
            {"dataset": "mnist", "F_type": "krr", "G_type": "linear", "score_mean": 0.2},
            {"dataset": "har", "F_type": "umap", "G_type": "linear", "score_mean": 0.3},
        ]
    ).to_csv(tmp_path / "result_grid_mnist.csv", index=False)

    output_path = export_result_grid_score_excel(tmp_path)
    ws = load_workbook(output_path).active

    assert ws["A1"].value == "dataset"
    assert ws["A2"].value == "G_type"
    assert ws["B1"].value == "mnist"
    assert ws["B2"].value == "umap"
    assert ws["C2"].value == "krr"
    assert ws["D1"].value == "har"
    assert ws["D2"].value == "umap"
    assert ws["A3"].value == "linear"
    assert ws["B3"].value == 0.1
