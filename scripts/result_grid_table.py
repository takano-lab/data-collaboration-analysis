from __future__ import annotations

from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = {"dataset", "F_type", "G_type", "score_mean"}


def list_result_grid_csvs(input_dir: str | Path) -> list[Path]:
    """Return result_grid_*.csv files, excluding result_grid_all.csv."""
    directory = Path(input_dir).expanduser()
    if not directory.is_dir():
        raise NotADirectoryError(f"input_dir is not a directory: {directory}")

    return sorted(
        path
        for path in directory.glob("result_grid_*.csv")
        if path.name != "result_grid_all.csv" and path.is_file()
    )


def load_result_grid_csvs(input_dir: str | Path) -> pd.DataFrame:
    """Load and concatenate result_grid_*.csv files from a directory."""
    csv_paths = list_result_grid_csvs(input_dir)
    if not csv_paths:
        raise FileNotFoundError(f"No result_grid_*.csv files found in {Path(input_dir).expanduser()}")

    frames: list[pd.DataFrame] = []
    missing_by_file: dict[str, list[str]] = {}

    for csv_path in csv_paths:
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        missing = sorted(REQUIRED_COLUMNS - set(df.columns))
        if missing:
            missing_by_file[str(csv_path)] = missing
            continue
        frames.append(df.assign(_source_file=csv_path.name))

    if missing_by_file:
        details = "; ".join(f"{path}: {cols}" for path, cols in missing_by_file.items())
        raise ValueError(f"Missing required columns: {details}")

    return pd.concat(frames, ignore_index=True)


def build_result_grid_score_table(
    input_dir: str | Path,
    *,
    aggfunc: str | Callable[[Iterable[float]], float] = "mean",
) -> pd.DataFrame:
    """
    Build a score_mean table with rows=G_type and columns=(dataset, F_type).

    The returned DataFrame has a two-level column MultiIndex:
    level 0 is dataset, level 1 is F_type.
    """
    df = load_result_grid_csvs(input_dir)
    table = df.pivot_table(
        index="G_type",
        columns=["dataset", "F_type"],
        values="score_mean",
        aggfunc=aggfunc,
        sort=False,
    )
    table.index.name = "G_type"
    table.columns.names = ["dataset", "F_type"]
    return table


def export_result_grid_score_excel(
    input_dir: str | Path,
    output_path: str | Path | None = None,
    *,
    sheet_name: str = "score_mean",
    aggfunc: str | Callable[[Iterable[float]], float] = "mean",
    number_format: str = "0.0000",
) -> Path:
    """
    Export score_mean table to Excel.

    Layout:
      - row 1: dataset labels
      - row 2: F_type labels
      - column A: G_type labels
    """
    input_dir = Path(input_dir).expanduser()
    if output_path is None:
        output_path = input_dir / "result_grid_score_mean_table.xlsx"
    output_path = Path(output_path).expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    table = build_result_grid_score_table(input_dir, aggfunc=aggfunc)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF4DE")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="dataset")
    ws.cell(row=2, column=1, value="G_type")
    ws.cell(row=1, column=1).font = bold
    ws.cell(row=2, column=1).font = bold
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=2, column=1).fill = subheader_fill

    for col_offset, (dataset, f_type) in enumerate(table.columns, start=2):
        ws.cell(row=1, column=col_offset, value=dataset)
        ws.cell(row=2, column=col_offset, value=f_type)
        for row in (1, 2):
            cell = ws.cell(row=row, column=col_offset)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill if row == 1 else subheader_fill

    start_col = 2
    while start_col <= table.shape[1] + 1:
        dataset = ws.cell(row=1, column=start_col).value
        end_col = start_col
        while end_col + 1 <= table.shape[1] + 1 and ws.cell(row=1, column=end_col + 1).value == dataset:
            end_col += 1
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=start_col).alignment = center
        start_col = end_col + 1

    for row_offset, (g_type, row) in enumerate(table.iterrows(), start=3):
        ws.cell(row=row_offset, column=1, value=g_type)
        ws.cell(row=row_offset, column=1).font = bold
        for col_offset, value in enumerate(row, start=2):
            cell = ws.cell(row=row_offset, column=col_offset, value=None if pd.isna(value) else float(value))
            cell.number_format = number_format

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = max(
        12,
        min(45, max(len(str(value)) for value in ["G_type", *table.index.tolist()]) + 2),
    )
    for col_idx in range(2, table.shape[1] + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(output_path)
    return output_path

